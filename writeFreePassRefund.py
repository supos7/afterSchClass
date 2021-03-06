#-*- coding: utf-8 -*-
# writeFreePassRefund.py

import logging
import os.path, shutil
import glob
import sqlite3
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import getopt, sys

# delete the previous log file
logfile = 'writeFreePassRefund.log'

FORMAT = "%(asctime)-15s  %(levelname)s %(message)s"
logging.basicConfig(filename=logfile, level=logging.DEBUG, format=FORMAT)

logging.info('<<<<< The log file of writeFreePassRefund.exe >>>>>')

def usage():
   print("Usage: writeFreePassRefund -y year -m month -f xls_folder -o out_folder")

try:
   opts, args = getopt.getopt(sys.argv[1:], "y:m:f:o:")
except getopt.GetoptError, err:
   # print help information and exit:
   print str(err) # will print something like "option -a not recognized"
   usage()
   os._exit(1)
if 4 != len(opts):
   usage()
   os._exit(1)
for o, a in opts:
   if o == "-y":
      cYear = a
   elif o == "-m":
      cMonth = a
   elif o == "-f":
      xlsDir = a
   elif o == "-o":
      outDir = a
   else:
      assert False, "unhandled option"

# open the afterSchool database
dbName = 'asClass.db'
if not os.path.isfile(dbName):
   logging.error("The database file \'asClass.db\' not found.")
   exit(1)

try:
   db = sqlite3.connect(dbName)
   cur = db.cursor()

   xlsPath = os.path.join(xlsDir,'*.xlsx')
   xlList = glob.glob(xlsPath)
   for xlFile in xlList:
      xlsName = xlFile[len(xlsDir)+1:].decode('cp949')
      logging.info(u'')
      logging.info(u'<<< ' + xlsName + u' >>>')
      wb = load_workbook(filename=xlFile)
      #for sheet in wb:
      sheet = wb.active
      logging.info(u'< ' + sheet.title + u' >')
      bBreak = False
      for row in sheet.rows:
         for cell in row:
            if cell.value and -1 < cell.value.find(u'학년'):
               rowFirst = cell.row +1
               colFirst = cell.col_idx -1 # 0 based
               bBreak = True
               break
         if bBreak:
            break
      if bBreak:
         # class
         className = xlsName[:xlsName.find(u'(')]
         t = (className, cYear, cMonth)
         cur.execute("SELECT id FROM afterSchoolClass WHERE cname = ? AND year = ? AND month = ?", t)
         r = cur.fetchone()
         if r is not None:
            classId = r[0]
         else:
            logging.info('The class not found: %s', className)
            continue
         bTuition = -1 < xlsName.find(u'강사')

         rowIdx = ()
         for row in sheet.rows:
            if row[colFirst].row < rowFirst:
               logging.debug('this line skipped: %s,%s,%s,%s,%s', row[colFirst].value, row[colFirst +1].value, row[colFirst +2].value, \
                  row[colFirst +3].value, row[colFirst +4].value)
               continue
            # student
            if row[colFirst].value and row[colFirst +1].value and row[colFirst +2].value and row[colFirst +3].value:
               if type(row[colFirst].value) is unicode:
                  stuGrade = row[colFirst].value.replace(u'학년',u'').strip()
               else:
                  stuGrade = row[colFirst].value
               if type(row[colFirst +1].value) is unicode:
                  stuClass = row[colFirst +1].value.replace(u'반',u'').strip()
               else:
                  stuClass = row[colFirst +1].value
               # class of the student
               t = (classId,stuGrade,stuClass,row[colFirst +2].value,row[colFirst +3].value,'FP','Y')
               if bTuition:
                  cur.execute("SELECT stuId,tuition FROM afterSchStu WHERE classId=? AND grade=? AND class=? AND odr=? AND name=? AND code=? AND tuit_pay=?", t)
               else:
                  cur.execute("SELECT stuId,mcost FROM afterSchStu WHERE classId=? AND grade=? AND class=? AND odr=? AND name=? AND code=? AND mcos_pay=?", t)
               r = cur.fetchone()
               row[colFirst +4].value = 0
               if r is not None:
                  if r[1] is not None:
                     row[colFirst +5].value = -r[1]
                     rowIdx = rowIdx + (row[0].row -1,)
            else:
               logging.debug('Invalid data: %s,%s,%s,%s,%s', row[colFirst].value, row[colFirst +1].value, row[colFirst +2].value, \
                  row[colFirst +3].value, row[colFirst +4].value)

         # arrange rows
         #for i in range(0,len(rowIdx)):
         #   r = sheet.rows[rowFirst -1 + i]
         #   row = sheet.rows[rowIdx[i]]
         #   for c, cell in zip(r, row):
         #      c.value = cell.value
         # delete residual
         #for r in range(rowFirst -1 + len(rowIdx), len(sheet.rows)):
         #   row = sheet.rows[r]
         #   for cell in row:
         #      cell.value = None
         #      cell.border = Border()
         # copy
         shTitle = sheet.title
         sheet.title = 'temp'
         wsOut = wb.create_sheet()
         wsOut.title = shTitle
         rowIdx = (0,) + rowIdx  # add the first row
         for i in range(0,len(rowIdx)):
            row = sheet.rows[rowIdx[i]]
            wsOut.append(range(len(row)))
            #r = wsOut.rows[rowFirst -1 + i]
            r = wsOut.rows[i]
            for c, cell in zip(r, row):
               c.value = cell.value
               if cell.has_style:
                  c.font = cell.font
                  c.border = cell.border
                  c.fill = cell.fill
                  c.number_format = cell.number_format
                  c.protection = cell.protection
                  c.alignment = cell.alignment
         # delete sheet
         wb.remove_sheet(sheet)         

      else: # not found '학년'
         logging.info(u'Worksheet \'' + sheet.title + u'\' may not have any student.')

      outPath = os.path.join(outDir, xlFile[len(xlsDir)+1:])
      wb.save(outPath)
      logging.info(u'Saved: ' + outPath.decode('cp949'))

   cur.close()
   db.commit()
   db.close()

except:
   logging.exception('Got an exception on database handler')
   raise

logging.info("Writing refund of free pass classes done.")
