#-*- coding: utf-8 -*-
# arrangeFreePass.py

import logging
import os.path, shutil
#import glob
#import sqlite3
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import getopt, sys

# delete the previous log file
logfile = 'arrangeFreePass.log'

FORMAT = "%(asctime)-15s  %(levelname)s %(message)s"
logging.basicConfig(filename=logfile, level=logging.DEBUG, format=FORMAT)

logging.info('<<<<< The log file of arrangeFreePass.exe >>>>>')

def usage():
   print("Usage: arrangeFreePass -f temp_xls -o out_xls")

try:
   opts, args = getopt.getopt(sys.argv[1:], "f:o:")
except getopt.GetoptError, err:
   # print help information and exit:
   print str(err) # will print something like "option -a not recognized"
   usage()
   os._exit(1)
if 2 != len(opts):
   usage()
   os._exit(1)
for o, a in opts:
   if o == "-f":
      tmpxls = a
   elif o == "-o":
      outxls = a
   else:
      assert False, "unhandled option"

print(outxls)
# open the afterSchool database
#dbName = 'asClass.db'
#if not os.path.isfile(dbName):
#   logging.error("The database file \'asClass.db\' not found.")
#   exit(1)

try:
#   db = sqlite3.connect(dbName)
#   cur = db.cursor()

#   xlsPath = os.path.join(xlsDir,'*.xlsx')
#   xlList = glob.glob(xlsPath)
#   for xlFile in xlList:
#      xlsName = xlFile[len(xlsDir)+1:].decode('cp949')
   xlsName = tmpxls.decode('cp949')
   logging.info(u'')
   logging.info(u'<<< ' + xlsName + u' >>>')
#      wb = load_workbook(filename=xlFile)
   wb = load_workbook(filename=tmpxls)
   #for sheet in wb:
   sheet = wb.active
   logging.info(u'< ' + sheet.title + u' >')
   bBreak = False
   for row in sheet.rows:
      for cell in row:
         if cell.value and -1 < cell.value.find(u'학년'):
            rowFirst = cell.row +1
            colFirst = cell.col_idx -1 # 0 based
            bBreak = True
            break
      if bBreak:
         break
   if bBreak:
#      # class
#      className = xlsName[:xlsName.find(u'(')]
#      t = (month,)
#      cur.execute("SELECT DISTINCT month FROM afterSchoolClass WHERE id IN (SELECT classId FROM classStu WHERE month=?)", t)
#      r = cur.fetchone()
#      t = (className, cYear, r[0])
#      cur.execute("SELECT id FROM afterSchoolClass WHERE cname=? AND year=? AND month=?", t)
#      r = cur.fetchone()
#      if r is not None:
#         classId = r[0]
#      else:
#         logging.info('The class not found: %s', className)
#         continue
#      bTuition = -1 < xlsName.find(u'수강')

#      t = ('FPN',month,classId)
#      if bTuition:
#         cur.execute("SELECT grade,class,odr,name,tuition FROM afterSchStu WHERE code=? AND month=? AND classId=? AND 0 < tuition ORDER BY grade,class,odr", t)
#      else: # mcost
#         cur.execute("SELECT grade,class,odr,name,mcost FROM afterSchStu WHERE code=? AND month=? AND classId=? AND 0 < mcost ORDER BY grade,class,odr", t)
#      iRow = 1
#      for r in cur:
#         row = sheet.rows[iRow]
#         row[colFirst].value = str(r[0]) + u'학년'
#         row[colFirst +1].value = str(r[1]) + u'반'
#         row[colFirst +2].value = str(r[2])
#         row[colFirst +3].value = r[3]
#         row[colFirst +4].value = 0
#         row[colFirst +5].value = r[4]
#         row[colFirst +6].value = -r[4]
#         iRow += 1

      # copy
      shTitle = sheet.title
      sheet.title = 'temp'
      wsOut = wb.create_sheet()
      wsOut.title = shTitle
      iRow = 0
      i = -1
#      for i in range(0,iRow):
      oldName = ''
      while iRow < sheet.max_row -1:
         iRow += 1
         row = sheet.rows[iRow]
         if row[colFirst].value and row[colFirst +1].value and row[colFirst +2].value:
            if oldName != row[colFirst +3].value:
               wsOut.append(28*[0])
#         wsOut.append(range(len(row)))
               i += 1
               r = wsOut.rows[i]
               for j in range(4):
                  r[j].value = row[colFirst +j].value
            off = 4 + 2 * (row[colFirst +4].value -1)
            for j in range(2):
               r[off +j].value = row[colFirst +5 + j].value
            oldName = row[colFirst +3].value
#            for c, cell in zip(r, row):
#               c.value = cell.value
#               if cell.has_style:
#                  c.font = cell.font
#                  c.border = cell.border
#                  c.fill = cell.fill
#                  c.number_format = cell.number_format
#                  c.protection = cell.protection
#                  c.alignment = cell.alignment
         else:
            break
      # delete sheet
      wb.remove_sheet(sheet)         

   else: # not found '학년'
      logging.info(u'Worksheet \'' + sheet.title + u'\' may not have any student.')

#   outPath = os.path.join(outDir, xlFile[len(xlsDir)+1:])
   outPath = outxls
   wb.save(outPath)
   logging.info(u'Saved: ' + outPath.decode('cp949'))

#   cur.close()
#   db.commit()
#   db.close()

except:
   logging.exception('Got an exception on database handler')
   raise

logging.info("Arranging free pass total done.")
