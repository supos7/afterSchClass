#-*- coding: utf-8 -*-
# checkStuListxls.py

import logging
import os.path, shutil
import glob
import sqlite3
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import getopt, sys

# delete the previous log file
logfile = 'checkStuListxls.log'

FORMAT = "%(asctime)-15s  %(levelname)s %(message)s"
logging.basicConfig(filename=logfile, level=logging.DEBUG, format=FORMAT)

logging.info('<<<<< The log file of checkStuListxls.exe >>>>>')

def usage():
   print("Usage: checkStuListxls -y year -m month -f xls_folder")

try:
   opts, args = getopt.getopt(sys.argv[1:], "y:m:f:o:")
except getopt.GetoptError, err:
   # print help information and exit:
   print str(err) # will print something like "option -a not recognized"
   usage()
   os._exit(1)
if 3 != len(opts):
   usage()
   os._exit(1)
for o, a in opts:
   if o == "-y":
      cYear = a
   elif o == "-m":
      cMonth = a
   elif o == "-f":
      xlsDir = a
   else:
      assert False, "unhandled option"

# open the afterSchool database
dbName = 'asClass.db'
if not os.path.isfile(dbName):
   logging.error("The database file \'asClass.db\' not found.")
   exit(1)

try:
   db = sqlite3.connect(dbName)
   cur = db.cursor()

   xlsPath = os.path.join(xlsDir,'*.xlsx')
   xlList = glob.glob(xlsPath)
   for xlFile in xlList:
      xlsName = xlFile[len(xlsDir)+1:].decode('cp949')
      logging.info(u'')
      logging.info(u'<<< ' + xlsName + u' >>>')
      wb = load_workbook(filename=xlFile)
      sheet = wb.active
      #logging.info(u'< ' + sheet.title + u' >')
      bBreak = False
      for row in sheet.rows:
         for cell in row:
            if cell.value and -1 < cell.value.find(u'학년'):
               rowFirst = cell.row +1
               colFirst = cell.col_idx -1 # 0 based
               bBreak = True
               break
         if bBreak:
            break
      if bBreak:
         # class
         className = xlsName[:xlsName.find(u'(')]
         t = (className, cYear, cMonth)
         cur.execute("SELECT id FROM afterSchoolClass WHERE cname = ? AND year = ? AND month = ?", t)
         r = cur.fetchone()
         if r is not None:
            classId = r[0]
         else:
            logging.info('The class not found: %s', className)
            continue
         bTuition = -1 < xlsName.find(u'강사')

         # make a dictionary
         stuDic = {}
         for row in sheet.rows:
            if row[colFirst].row < rowFirst:
               logging.debug('this line skipped: %s,%s,%s,%s,%s', row[colFirst].value, row[colFirst +1].value, row[colFirst +2].value, \
                  row[colFirst +3].value, row[colFirst +4].value)
               continue
            # student
            if row[colFirst].value and row[colFirst +1].value and row[colFirst +2].value and row[colFirst +3].value:
               if type(row[colFirst].value) is unicode:
                  stuGrade = row[colFirst].value.replace(u'학년',u'').strip()
               else:
                  stuGrade = row[colFirst].value
               if type(row[colFirst +1].value) is unicode:
                  stuClass = row[colFirst +1].value.replace(u'반',u'').strip()
               else:
                  stuClass = row[colFirst +1].value
               
               key = u'%d%02d%02d%s' % (int(stuGrade),int(stuClass),int(row[colFirst +2].value),row[colFirst +3].value)
               stuDic[key] = 'OK'
            else:
               logging.debug('Invalid data: %s,%s,%s,%s,%s', row[colFirst].value, row[colFirst +1].value, row[colFirst +2].value, \
                  row[colFirst +3].value, row[colFirst +4].value)

         dbTitles = [u'방과후 행정사 파일', u'행정실 파일']
         t = (classId,)
         if bTuition:
            cur.execute("SELECT cname,grade,class,odr,name FROM afterSchStu WHERE classId=? AND tuition IS NOT NULL ORDER BY grade,class,odr", t)
         else:
            cur.execute("SELECT cname,grade,class,odr,name FROM afterSchStu WHERE classId=? AND mcost IS NOT NULL ORDER BY grade,class,odr", t)
         bf = True
         for row in cur:
            key = u'%d%02d%02d%s' % (int(row[1]),int(row[2]),int(row[3]),row[4])
            if stuDic.get(key) is None:
               if bf:
                  logging.info(dbTitles[0] + u'에는 있고, ' + dbTitles[1] + u'에는 없는 학생')
                  bf = False
               logging.info(u'%s: %s학년 %s반 %s번 %s', row[0],row[1],row[2],row[3],row[4])

         # swap
         dbTitles[0],dbTitles[1] = dbTitles[1],dbTitles[0]

         rowIdx = ()
         bf = True
         for row in sheet.rows:
            if row[0].row < rowFirst:
               continue
            # student
            if row[colFirst].value and row[colFirst +1].value and row[colFirst +2].value and row[colFirst +3].value:
               if type(row[colFirst].value) is unicode:
                  stuGrade = row[colFirst].value.replace(u'학년',u'').strip()
               else:
                  stuGrade = row[colFirst].value
               if type(row[colFirst +1].value) is unicode:
                  stuClass = row[colFirst +1].value.replace(u'반',u'').strip()
               else:
                  stuClass = row[colFirst +1].value
               
               t = (classId,stuGrade,stuClass,row[colFirst +2].value,row[colFirst +3].value)
               if bTuition:
                  cur.execute("SELECT stuId,cname FROM afterSchStu WHERE classId=? AND grade=? AND class=? AND odr=? AND name=? AND tuition IS NOT NULL", t)
               else:
                  cur.execute("SELECT stuId,cname FROM afterSchStu WHERE classId=? AND grade=? AND class=? AND odr=? AND name=? AND mcost IS NOT NULL", t)
               r = cur.fetchone()
               if r is None:
                  if bf:
                     logging.info(dbTitles[0] + u'에는 있고, ' + dbTitles[1] + u'에는 없는 학생')
                     bf = False
                  logging.info(u'%s: %d학년 %d반 %d번 %s', className,int(stuGrade),int(stuClass),int(row[colFirst +2].value),row[colFirst +3].value)

      else: # not found '학년'
         logging.info(u'Worksheet \'' + sheet.title + u'\' may not have any student.')

   cur.close()
   db.commit()
   db.close()

except:
   logging.exception('Got an exception on database handler')
   raise

logging.info("Checking student list between the database and xls files done")
